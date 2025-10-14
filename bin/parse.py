#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from operator import ne
import sys
import os
import gzip
import csv
from pathlib import Path
import traceback
from collections import Counter, OrderedDict
import itertools
import json
from copy import deepcopy
from io import StringIO, BytesIO
from shutil import copyfileobj
from string import ascii_uppercase
import functools
import chardet

import openpyxl
from pandas.core.base import NoNewAttributesMixin

try:
    import bin.util as util
except ModuleNotFoundError:
    import util

try:
    import bin.transaction as transaction
except ModuleNotFoundError:
    import transaction
    
from stutil import m

transact = transaction.transact
untransact = transaction.untransact


"""
@created: Nov 2021
@author: Cameron Jack, Bob Buckley, ANU Bioinformatics Consultancy, 2019-2021

Supports:
 * Hamilton Nimbus generate input and read output files
 * Labcyte Echo generate input files (picklists) for primer PCR reactions and index PCR reactions
 * Illumina Miseq generate samplesheet input files

Contains all file IO methods that involved barcodes, that are not purely for user reporting or robot inputs. 
Trivial file reads which don't involve barcodes should stay with the rest of their functionality for better readability
Incorporates FASTA checking and parsing code by Eslam Ibrahim, ANU Bioinformatics Consultancy, 2024

The following rules must be followed to protect users from typing mistakes and MS Excel weirdness:
* Barcodes should be guarded when first read from a new source
* Barcodes should be written with guards if the file is to be read again by the pipeline
* The pipeline should look for guards in any file that was written by the pipeline - for the pipeline
* Barcodes should be unguarded when written as machine inputs, or for final reporting
* All rows and fields should be stripped of white space immediately
* All file reads should be done with ignore=True to avoid non-ASCII characters
* All file reads should protect against empty rows
* Command-line interfaces and web interfaces expect unguarded barcodes but must be explicit for which type of barcode is needed
* The interface codes are responsible for providing guarded barcodes to all internal functions
* Internal functions only accept guarded barcodes

All uploaded files go through upload(exp, files, purpose) and then process_upload(exp, file, purpose) once
the file has been cleared as a safe transaction. After this the various parsers are called. Finally
we rename any transacted plates that are accepted by the user.

Important note: files are state and are immutable in the eyes of the pipeline. Plates are mutable and are
never held in Experiment.reproducible_steps at the top level, and they are never saved as files by the pipeline.

Behaves like a C++ "friend" of the Experiment class - very tightly coupled.

Our own FASTA parser and checking code is here, replacing BioPython's SeqIO.FastaIO.SimpleFastaParser
"""

def process_upload_queue(exp):
    """
    call upload() on each entry in the exp['file_uploads']['_upload_queue'] = {}
    """
    if '_upload_queue' not in exp.uploaded_files:
        exp.uploaded_files['_upload_queue'] = {}
    else:
        pfps_to_clear = []
        for pfp in exp.uploaded_files['_upload_queue']:
            purpose, caller_id, overwrite_plates = exp.uploaded_files['_upload_queue'][pfp]  
            upload(exp, pfp, purpose, caller_id=caller_id, overwrite_plates=overwrite_plates)
            pfps_to_clear.append(pfp)
        for pfp in pfps_to_clear:
            del exp.uploaded_files['_upload_queue'][pfp]
        #exp.uploaded_files['_upload_queue'] = {}
        #m(f'upload queue cleared', level='info', dest=('noGUI',))
    

### Universal upload interface

def upload(exp, pfp, purpose, caller_id=None, overwrite_plates=True, overwrite_files=False):
    """
    All file uploads go through this function, having been previously queued
    - Each file is already uploaded with a pending_ prefix in /exp/uploads/ (pfp - protected file path)
    - it is added to exp.uploaded_files['_upload_pending'] = {}
    - or is sent to process_upload(exp, file, purpose)
    
    Purpose is required for the upload, caller_id is the display unit associated with the upload
    
    Return True if uploaded stream is correctly saved and parsed, else False
    By default all entries are overwritten rather than protected - this is a practical issue for NGS Genotyping
    """
    if exp.locked and purpose not in {'primer_assay_map','rodentity_reference','amplicon_reference'}:
        m(f'Cannot add files other than assay file or references while lock is active', level='failure', caller_id=caller_id)
        return False

    if '_upload_pending' not in exp.uploaded_files:
        exp.uploaded_files['_upload_pending'] = {}

    final_fp = transaction.untransact(pfp) 
    if Path(final_fp).exists():
        if overwrite_files:
            s = util.delete_file(final_fp)
            if not s:
                m(f'could not overwrite original file {final_fp}', level='error', caller_id=caller_id)
                exp.uploaded_files['_upload_pending'][pfp] = (purpose, caller_id, overwrite_plates)
                return False
            s = process_upload(exp, pfp, purpose, caller_id=caller_id, overwrite_plates=overwrite_plates)
            if s is False:
                exp.del_file_record(final_fp)
                util.delete_file(final_fp)
                m(f'could not upload file {final_fp}', level='error', caller_id=caller_id)
                return False
            else:
                m(f'uploaded file {final_fp}', level='success', caller_id=caller_id)
                return True
        else:
            # the user will decide later whether to overwrite what they have
            m(f'adding {pfp} to pending uploads', level='info', dest=('noGUI',))
            exp.uploaded_files['_upload_pending'][pfp] = (purpose, caller_id, overwrite_plates)
            return True
    else:
        
        s = process_upload(exp, pfp, purpose, caller_id=caller_id, overwrite_plates=overwrite_plates)
        if s is False:
            exp.del_file_record(final_fp)
            util.delete_file(final_fp)
            m(f'could not upload file {final_fp}', level='error', caller_id=caller_id)
            return False
        else:
            m(f'uploaded file {final_fp}', level='success', caller_id=caller_id)
            return True


def accept_pending_upload(exp, pfp, purpose, caller_id=None, overwrite_plates=True):
    """
    Take a pending upload file pfp, its purpose, and attempt to parse the file
    """
    if exp.locked and purpose not in {'primer_assay_map','rodentity_reference','amplicon_reference'}:
        m('Cannot add manifest while lock is active', level='failure', caller_id=caller_id)
        return False

    final_fn = transaction.untransact(pfp)
    if Path(final_fn).exists():  # it should
        if final_fn in exp.uploaded_files:
            success = exp.del_file_record(final_fn, soft=True, caller_id=caller_id)
        else:
            success = util.delete_file(final_fn, caller_id=caller_id)
        if not success:
            m(f'could not delete original file {final_fn}', level='error', caller_id=caller_id)
            return False
    
    if pfp in exp.uploaded_files['_upload_pending']:
        del exp.uploaded_files['_upload_pending'][pfp]
    success = process_upload(exp, pfp, purpose, caller_id=caller_id, overwrite_plates=overwrite_plates)
    if not success:
        exp.del_file_record(final_fn)
        
    return True
        

def process_upload(exp, pfp, purpose, caller_id=None, overwrite_plates=True):
    """
    Called by upload() or accept_pending_upload()
    - assumes no impediments and renames file to final path name
    - calls exp.add_file_record(filepath, purpose)
    - invokes the appropriate parser for the provided purpose and file extension
    purposes: ['amplicon','DNA','pcr','rodentity_sample','custom_sample','primer_layout','primer_volume',
            'index_layout','index_volume','primer_assay_map','rodentity_reference','amplicon_reference','taq_water']
    plates now get loaded with {filepath:purpose}
            
    caller_id (str) is the name of the associated display unit for user messages
    
    Saves the experiment on successful upload and parse
    """
    fp = transaction.untransact(pfp)
    if not Path(pfp).exists():
        m(f'{pfp} not longer exists for upload, cancelling', level='warning', dest=('noGUI',))
        return False
    
    if Path(fp).exists():
        try:
            Path.unlink(fp)
        except Exception as exc:
            m(f'failed to remove existing file {fp}', level='error', caller_id=caller_id)
            return False
    try:
        Path(pfp).rename(fp)
    except Exception as exc:
        m(f'failed to rename {pfp} to {fp}, {exc}', level='error', caller_id=caller_id)
        util.delete_file(pfp)
        return False    

    exp.add_file_record(fp, purpose=purpose)
    pids = None  # list of plateIDs associated with a file
    if purpose == 'amplicon':
        success, pids = parse_amplicon_manifest(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'DNA' or purpose == 'dna':  # Echo_COC file
        success, pids = parse_dna_plate(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'rodentity_sample':  # becomes purpose=sample source=rodentity
        success, pids = parse_rodentity_json(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'custom_sample':  # becomes purpose=sample source=custom
        success, pids = parse_custom_manifest(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'custom_384_sample':  # actually 'dna' plates in this system
        success, pids = parse_custom_manifest(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates,
                pformat='Echo384')
    elif purpose == 'pcr':
        m('no parser implemented for PCR plate records', level='critical', caller_id=caller_id)
        # parse_pcr_record(exp, fp)
    elif purpose == 'primer_layout':
        success, pids = parse_primer_layout(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'primer_volume':
        success, pids = parse_primer_volume(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'index_layout':
        success, pids = parse_index_layout(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'index_volume':
        success, pids = parse_index_volume(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'assay_primer_map':
        success = parse_assay_primer_map(exp, fp, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose in {'rodentity_reference', 'custom_reference','amplicon_reference'}:
        success = parse_reference_sequences(exp, fp, purpose, caller_id=caller_id, overwrite_plates=overwrite_plates)
    elif purpose == 'taq_water':
        m('no parser implemented for taq_water plates', level='critical', caller_id=caller_id)
    else:
        m(f'no parser implemented for {purpose=}', level='critical', caller_id=caller_id)

    if success:
        #m(f'parsed {fp} for {purpose}', level='success', caller_id=caller_id)
        exp.mod_file_record(fp, extra_PIDs=pids)
        exp.save()
        return True
    else:
        #m(f'could not parse {fp} for {purpose}', level='error', caller_id=caller_id)
        exp.del_file_record(fp, soft=True)
        return False


### parsers that build (mostly) json-like dictionary structures for Experiment

# parse_rodentity_json
# parse_custom_manifest
# parse_primer_assay_mapping
# parse_reference_sequences
# parse_taqwater - not necessary, generated
# parse_targets - use parse_reference_sequences
# parse_primer_layout
# parse_primer_volume
# parse_index_layout
# parse_index_volume
# parse_amplicons

#@functools.lru_cache
#def load_json(fp):
#    """ JSON data loader that is cached for fast reading """
#    with open(fp, 'rt', errors='ignore') as src:
#        info = json.load(src) 
        

def parse_rodentity_json(exp, fp, caller_id=None, overwrite_plates=True):
    """
    Read a JSON file containing one or more Rodentity plate definitions 
    Write a standardised datastructure into exp.plate_location_sample
    Update entry in exp.uploaded_files to include plate ids
    
    caller_id (str) is the name of the associated display unit for user messages
    """
    m(f'parsing Rodentity JSON file {fp}', level='begin')
    well_records = {}  # [gpid][pos]
    with open(fp, 'rt', errors='ignore') as src:
        info = json.load(src)  

    for record in info['wells']:
        pid = record['plate']
        gpid = util.guard_pbc(pid, silent=True)
        if gpid not in well_records:
            well_records[gpid] = {'purpose':'sample','source':'rodentity', 'wells':set(), 'plate_type':'96'}
            
        pos = util.unpadwell(record['wellLocation'])
        if pos in well_records[gpid] and well_records[gpid][pos]['wells'] != {}:
            # duplicate entry for a given plate and well position
            m(f'skipping duplicate sample entry {pid} {pos} in {fp}', level='warning', caller_id=caller_id)
            continue
        well_records[gpid]['wells'].add(pos)
        well_records[gpid][pos] = {}
        well_records[gpid][pos]['barcode'] = util.guard_rbc(record['mouse']['barcode'], silent=True)
        well_records[gpid][pos]['sampleBarcode'] = well_records[gpid][pos]['barcode']  # as the line above
        well_records[gpid][pos]['samplePlate'] = gpid
        well_records[gpid][pos]['sampleWell'] = pos
        well_records[gpid][pos]['strain'] = record['mouse']['mouselineName']
        well_records[gpid][pos]['sex'] = record['mouse']['sex']
        well_records[gpid][pos]['mouse'] = deepcopy(record['mouse'])
            
        # build all the various assay/allele structures required
        well_records[gpid][pos]['ngs_assay_records'] = {}
        well_records[gpid][pos]['other_assay_records'] = {} 
        ngs_assays = []
        other_assays = []
                    
        if 'alleles' in record['mouse']:
            for allele in record['mouse']['alleles']:  # allele is a dict from a list
                for assay in allele['assays']:  # assay is a dict from a list
                    assay_name = str(assay['name'])
                    assay_method = str(assay['method'])
                    assay_key = str(assay['assay_key'])
                    allele_key = str(allele['alleleKey'])
                    allele_symbol = '"'+str(allele['symbol'])+'"'
                    if assay_method == 'NGS' or assay_name.startswith('NGS'):
                        ngs_assays.append(assay_name)
                        ngs_or_other = 'ngs_assay_records'
                    else:
                        other_assays.append(assay_name)
                        ngs_or_other = 'other_assay_records'
                    if assay_name not in well_records[gpid][pos][ngs_or_other]:
                        well_records[gpid][pos][ngs_or_other][assay_name] = {}
                    well_records[gpid][pos][ngs_or_other][assay_name]['alleleKey'] = allele_key
                    well_records[gpid][pos][ngs_or_other][assay_name]['alleleSymbol'] = allele_symbol
                    well_records[gpid][pos][ngs_or_other][assay_name]['assayKey'] = assay_key
                    well_records[gpid][pos][ngs_or_other][assay_name]['assayName'] = assay_name # redundant
                    well_records[gpid][pos][ngs_or_other][assay_name]['assayMethod'] = assay_method # not entirely redundant
                            
        well_records[gpid][pos]['ngs_assays'] = ngs_assays.copy()
        well_records[gpid][pos]['other_assays'] = other_assays.copy()
    
    for gpid in well_records:
        if gpid in exp.plate_location_sample:
            if not overwrite_plates:
                m(f'overwriting disallowed for existing plate entry {gpid}', level='failure', caller_id=caller_id)
                continue
            elif exp.plate_location_sample[gpid]['purpose'] != 'sample':
                m(f'plate {gpid} already exists with purpose {exp.plate_location_sample[gpid]["purpose"]}', 
                        level='error', caller_id=caller_id)
                continue
            else:
                exp.plate_location_sample[gpid] = {}  # wipe the existing entry
            well_records[gpid]['filepath_purpose'] = {fp:'rodentity_sample'}
        exp.plate_location_sample[gpid] = well_records[gpid].copy()

    m(f'parsing {fp} with plates: {list(well_records.keys())}', level='end')
    return True, list(well_records.keys())
            

def parse_custom_manifest(exp, fp, caller_id=None, overwrite_plates=True,
        pformat='96'):
    """
    Parse a custom manifest files and store them in self.unassigned_plates['custom'] =\
            {plateBarcode={Assay=[],...}}
    Also add an entry for each file into self.uploaded_files
    Example headers:
    sampleNum	plateBarcode	well	sampleBarcode	assay	assay	clientName	sampleName 	alleleSymbol
    Required columns: [plateBarcode, well, sampleBarcode, assay*, clientName] *Duplicates allowed
    Optional columns: [sampleNo, sampleName, alleleSymbol] and anything else
    
    caller_id (str) is the name of the associated display unit for user messages
    pformat is either '96' or 'Echo384', both defined in util.py
    """
    m(f'Parsing custom manifest {fp}', level='begin')
    if fp.lower().endswith('xlsx'):
        with open(fp, 'rb') as bytes:
            workbook = openpyxl.load_workbook(bytes)
            sheet = workbook.active
            file_table = [','.join(map(str,cells)) for cells in sheet.iter_rows(values_only=True)]
    else:
        with open(fp, 'rt', errors='ignore') as src:
            file_table = [line.strip() for line in src]

    well_records = {}  # [plate] = {well:{record}}
    plate_barcode_col = None
    assay_cols = []
    required_cols = ['platebarcode', 'well', 'samplebarcode', 'assay', 'clientname'] 
    for i, line in enumerate(file_table):
        cols = [c.strip() if c is not None else '' for c in line.split(',')]  # lower case column names
        #print(i, cols, file=sys.stderr)
        if i==0:  # process header
            cols_lower = [c.lower() for c in cols]
            header_dict = {k:c for k,c in enumerate(cols)}
            #print(header_dict, file=sys.stderr)
            matching_cols = [col_name in cols_lower for col_name in required_cols]
            if not all(matching_cols):
                m(f'manifest {fp} requires at least columns {required_cols}', level='error', caller_id=caller_id)
                return False, []
            
            for k in header_dict:
                if header_dict[k].lower() == 'platebarcode':
                    plate_barcode_col = k
                elif header_dict[k].lower() == 'assay':
                    assay_cols.append(k)
            continue
            
        if len(cols) < 5:  # skip empty rows
            continue
        pid = cols[plate_barcode_col]
        gpid = util.guard_pbc(pid, silent=True)
        
        if gpid not in well_records and pformat == '96':
            well_records[gpid] = {'purpose':'sample','source':'custom', 'wells':set(),
                    'plate_type':util.PLATE_TYPES[pformat]}
        elif gpid not in well_records and pformat == 'Echo384':
            well_records[gpid] = {'purpose':'dna','source':'custom', 'wells':set(),
                    'plate_type':util.PLATE_TYPES[pformat]}
            
        assays = [] 
        # do a first pass to set up recording a sample in a well
        for k,c in enumerate(cols):
            if c.lower() == 'none':
                continue
            if k in assay_cols:
                if c != '':  # ignore empty assay entries
                    assays.append(c)
            if header_dict[k].lower() == 'well':
                well = util.unpadwell(c.upper())
                #print(f'well: {well=} {k=} {c=}')
                if well in well_records[gpid]['wells'] and well_records[gpid][well] != {}:
                    m(f"duplicate {c} in {pid}. Skipping row {i+2} {cols=}", level='warning', caller_id=caller_id)
                    continue
                well_records[gpid]['wells'].add(well)
                well_records[gpid][well] = {'sampleWell':well}
        
        # now do a second pass to collect everything together
        for k,c in enumerate(cols):
            if c.lower() == 'none':
                c = ''
            if header_dict[k].lower() == 'well':
                continue
            if header_dict[k].lower() == 'samplebarcode':
                #if c.startswith('C'):
                #    sid = util.guard_cbc(c, silent=True)
                #elif c.startswith('M'):
                #    sid = util.guard_rbc(c, silent=True)
                #else:
                sid = util.guard_cbc(c, silent=True) # fall back to custom?
                well_records[gpid][well]['barcode'] = sid
                well_records[gpid][well]['sampleBarcode'] = sid
            elif header_dict[k].lower() == 'platebarcode':
                well_records[gpid][well]['platebarcode'] = gpid
                well_records[gpid][well]['samplePlate'] = gpid
            elif header_dict[k].lower() == 'sampleno':
                well_records[gpid][well]['sampleNumber'] = str(c)
            elif header_dict[k].lower() == 'assay':
                continue
            elif header_dict[k].lower() == 'samplename':
                well_records[gpid][well]['sampleName'] = str(c)
            elif header_dict[k].lower() == 'clientname':
                well_records[gpid][well]['clientName'] = str(c)
            else:
                try:
                    well_records[gpid][well][header_dict[k]] = str(c)
                except:
                    m(f'Failed to record well_record {gpid=} {well=} {header_dict[k]=} {c=}',
                            level='critical', caller_id=caller_id)
        well_records[gpid][well]['ngs_assays'] = assays

    for gpid in well_records:
        if gpid in exp.plate_location_sample:
            if not overwrite_plates:
                m(f'Overwriting turned off, cannot overwrite existing plate entry {gpid}', 
                        level='failure', caller_id=caller_id)
                continue
            elif pformat == '96' and exp.plate_location_sample[gpid]['purpose'] != 'sample':
                m(f'plate {gpid} already exists with purpose {exp.plate_location_sample[gpid]["purpose"]}', 
                        level='error', caller_id=caller_id)
                continue
            elif pformat == 'Echo384' and exp.plate_location_sample[gpid]['purpose'] != 'dna':
                m(f'plate {gpid} already exists with purpose {exp.plate_location_sample[gpid]["purpose"]}', 
                        level='error', caller_id=caller_id)
                continue
            else:
                exp.plate_location_sample[gpid] = {}  # create empty entry
        if pformat == '96':
            well_records[gpid]['filepath_purpose'] = {fp:'custom_sample'}
        elif pformat == 'Echo384':
            well_records[gpid]['filepath_purpose'] = {fp:'custom_384'}
        exp.plate_location_sample[gpid] = well_records[gpid].copy()

    m(f'parsing {fp} with plates: {list(well_records.keys())}', level='end')
    return True, list(well_records.keys())


def parse_amplicon_manifest(exp, fp, caller_id=None, overwrite_plates=True):
    """
    Amplicon plate layouts (containing pre-amplified sequences) are parsed here from manifest files.
    Only column layout (comma separate) is supported. col1: plate barcode; col2: well position; col3: sample barcode; col4 (optional): volume.
    The first row must be a header row: plate, well, sample, (volume in uL, if provided).
    plateBarcode, well, sampleBarcode, amplicon* (*multiple columns with this name are allowed)
    Adding this will result in changing the Miseq and Stage3 output files, so needs to be wrapped as a transaction
    Needs to check whether there are sufficient indexes
    
    caller_id (str) is the name of the associated display unit for user messages
    """
    m(f'parsing amplicon manifest {fp}', level='begin')
    if fp.lower().endswith('xlsx'):
        with open(fp, 'rb') as bytes:
            workbook = openpyxl.load_workbook(bytes)
            sheet = workbook.active
            file_table = [','.join(map(str,cells)) for cells in sheet.iter_rows(values_only=True)]
    else:
        with open(fp, 'rt', errors='ignore') as src:
            file_table = [line.strip() for line in src]

    well_records = {}     
    plate_barcode_col = None
    well_col = None
    amplicon_cols = []  # combine these
    required_cols = ['platebarcode', 'well', 'samplebarcode']
    for i, line in enumerate(file_table): # csv_reader is super limiting
        cols = [c.strip() if c is not None else '' for c in line.split(',')]  # lower case column names
        if i==0:  # process header
            cols_lower = [c.lower() for c in cols]
            header_dict = {k:c for k,c in enumerate(cols_lower)}
            matching_cols = [col_name in cols_lower for col_name in required_cols] 
            if not all(matching_cols):
                m(f'amplicon manifest {fp} requires at least {required_cols} columns', 
                        level='error', caller_id=caller_id)
                return False, []
            
            for k in header_dict:
                if header_dict[k] == 'platebarcode':
                    plate_barcode_col = k
                elif header_dict[k] == 'well':
                    well_col = k
                elif header_dict[k] == 'amplicon': # combine amplicon columns
                    amplicon_cols.append(k)
            continue
        if len(cols) < 3:  # skip empty rows
            continue
        # set up plate
        pid = cols[plate_barcode_col]
        gpid = util.guard_pbc(pid, silent=True)
        if gpid not in well_records:
            well_records[gpid] = {'purpose':'amplicon','source':'custom', 'wells':set(), 
                    'plate_type':'384PP_AQ_BP'}

        # set up well
        pos = util.unpadwell(cols[well_col].upper())
        if pos in well_records[gpid]['wells'] and well_records[gpid][pos] != {}:
            m(f"duplicate {c} in {pid}. Skipping row {i+2} {cols=}", level='warning', caller_id=caller_id)
            continue
        well_records[gpid]['wells'].add(pos)
        well_records[gpid][pos] = {'sampleWell':pos}
        amplicons = []
        # now collect everything together
        for k,c in enumerate(cols):
            if c.lower() == 'none':
                c = ''
            if header_dict[k] == 'well':
                continue  # we've already got this!
            if header_dict[k] == 'samplebarcode':
                # amplicon guards
                sid = well_records[gpid][pos][header_dict[k]] = util.guard_abc(c, silent=True)
                well_records[gpid][pos]['barcode'] = sid
                well_records[gpid][pos]['sampleBarcode'] = sid
            elif header_dict[k] == 'platebarcode':
                # don't really need it but whatever - record the guarded PID
                well_records[gpid][pos]['platebarcode'] = gpid
                well_records[gpid][pos]['samplePlate'] = gpid
            elif header_dict[k] == 'sampleno':
                well_records[gpid][pos]['sampleNumber'] = str(c)
            elif header_dict[k] == 'amplicon':
                if c != '':  # ignore empty amplicon entries
                    amplicons.append('"'+c+'"')
            elif header_dict[k] == 'volume':
                well_records[gpid][pos]['volume'] = float(c)*1000  # save as nL
            else:
                try:
                    well_records[gpid][pos][header_dict[k]] = str(c)
                except:
                    print(type(header_dict[k]), header_dict[k], str(c), file=sys.stderr)
        well_records[gpid][pos]['amplicons'] = amplicons

    for gpid in well_records:
        if gpid in exp.plate_location_sample:
            if not overwrite_plates:
                m(f'Overwriting plates disabled, cannot overwrite existing plate entry {gpid}', 
                        level='failure', caller_id=caller_id)
                continue
            elif exp.plate_location_sample[gpid]['purpose'] != 'amplicon':
                m(f'plate {gpid} already exists with purpose {exp.plate_location_sample[gpid]["purpose"]}', 
                        level='error', caller_id=caller_id)
                continue
        well_records[gpid]['filepath_purpose'] = {fp:'amplicon'}
        exp.plate_location_sample[gpid] = well_records[gpid].copy()

    m(f'parsing {fp} with plates: {list(well_records.keys())}', level='end')
    return True, list(well_records.keys())


def parse_primer_layout(exp, fp, user_gpid=None, caller_id=None, overwrite_plates=True):
    """
    add primer plate definition with well and name columns
    caller_id (str) is the name of the associated display unit for user messages
    """
    if user_gpid:
        PID = util.unguard_pbc(user_gpid, silent=True)
        gPID = util.guard_pbc(user_gpid, silent=True)
    else:
        PID = Path(fp).name.split('_')[0]  # assumes first field is PID
        gPID = util.guard_pbc(PID, silent=True)

    # parse the file first to make sure it works
    well_records = {gPID:{}}
    with open(fp, 'rt', errors='ignore') as data:
        for i, row in enumerate(csv.reader(data, delimiter=',', quoting=csv.QUOTE_MINIMAL)):
            if i == 0:
                continue  # header
            if row == '' or row[0] == '' or row[1] == '':
                continue
            try:
                well = util.unpadwell(row[0])
            except UnboundLocalError as exc:
                m(f'You likely tried to upload volumes instead of layout! Using file {fp} {exc}',
                        level='error', caller_id=caller_id)
                return False, []
            if well in well_records[gPID]:
                m(f'skipping duplicate well entry {well} in {fp}', level='warning', caller_id=caller_id)
                continue 
            well_records[gPID][well] = {'primer':row[1]}
           
    # check if the plate already exists and that its purpose is related but not radically different
    if gPID in exp.plate_location_sample:
        plate_purpose = exp.plate_location_sample[gPID].get('purpose', None) 
        if plate_purpose is None:
            # this should not happen, every plate has a purpose. Wipe the existing plate
            exp.delete_plate(gPID)
            exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                    'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_layout'}}
            m(f"Creating new primer plate record for {PID}", level='info', caller_id=caller_id)
        elif plate_purpose != 'primer':
            m(f"Primer plate PID: {PID} matches existing plate entry of different purpose "+\
                    f"{exp.plate_location_sample[gPID]['purpose']}", level='error', caller_id=caller_id)
            return False, []
        
        # now check that the filepaths make sense
        filepath_purpose = exp.plate_location_sample[gPID].get('filepath_purpose', None)
        if not isinstance(filepath_purpose, dict) or len(filepath_purpose) == 0:
            exp.plate_location_sample[gPID]['filepath_purpose'] = {fp:'primer_layout'}
        else:
            if fp in filepath_purpose:
                if filepath_purpose[fp] == 'primer_layout':
                    if not overwrite_plates:
                        m(f'Primer layout plate already found in experiment, overwriting not allowed', 
                                level='failure', caller_id=caller_id)
                        return False, []
                    m(f'replacing existing primer layout for {gPID}, wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_layout'}}
                      
                elif filepath_purpose[fp] not in ['primer_layout', 'primer_volume']:
                    m(f'existing file of path {fp} has new purpose "primer_layout" instead of {filepath_purpose[fp]}, '+ 
                            f'wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_layout'}}
    else:
        exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_layout'}}
                      
    # copy across data from the temporary plate records                           
    for well in well_records[gPID]:
        if well not in exp.plate_location_sample[gPID]:
            exp.plate_location_sample[gPID]['wells'].add(well)
            exp.plate_location_sample[gPID][well] = {}
        exp.plate_location_sample[gPID][well]['primer'] = well_records[gPID][well]['primer']
                    
    m(f"Done with primer layout from {fp}", level='end')
    return True, [gPID]


def parse_primer_volume(exp, fp, user_gpid=None, caller_id=None, overwrite_plates=True):
    """
    add primer plate volumes from either plate layout or two-column layout
    caller_id (str) is the name of the associated display unit for user messages
    """
    if user_gpid:
        PID = util.unguard_pbc(user_gpid, silent=True)
        gPID = util.guard_pbc(user_gpid, silent=True)
    else:
        PID = Path(fp).name.split('_')[0]  # assumes first field is PID
        gPID = util.guard_pbc(PID, silent=True)

    # parse the file first to make sure it works
    well_records = {gPID:{}}
    with open(fp, 'rt', errors='ignore') as data:
        for i, row in enumerate(csv.reader(data, delimiter=',', quoting=csv.QUOTE_MINIMAL)):
            if i == 0:
                if row[0].lower().startswith('date'):
                    # plate style layout
                    layout = 'plate'
                else:
                    layout = 'columns'
                continue  # header
                
            if layout == 'plate':
                if not row or row[0] == '' or row[0][0] not in ascii_uppercase[:16]:  # A-P
                    continue
                for j,col in enumerate(row):
                    if j==0:
                        continue  # row name cell
                    well = util.unpadwell(row[0] + str(j))
                    if well in well_records[gPID]:
                        m(f'skipping duplicate well entry {well} in {fp}', level='warning', dest=('noGUI',))
                        continue
                    well_records[gPID][well] = {'volume':float(col)*1000}
            else:  # columns
                well = util.unpadwell(row[0])
                if well in well_records[gPID]:
                    m(f'skipping duplicate well entry {well} in {fp}', level='warning', dest=('noGUI',))
                    continue
                well_records[gPID][well] = {'volume':float(col)*1000}

    # check if the plate already exists and that its purpose is related but not radically different
    if gPID in exp.plate_location_sample:
        plate_purpose = exp.plate_location_sample[gPID].get('purpose', None) 
        if plate_purpose is None:
            # this should not happen, every plate has a purpose. Wipe the existing plate
            exp.delete_plate(gPID)
            exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                    'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_volume'}}
            m(f"Creating new primer plate record for {PID}", level='info', caller_id=caller_id)
        elif plate_purpose != 'primer':
            m(f"Primer plate PID: {PID} matches existing plate entry of different purpose "+\
                    f"{exp.plate_location_sample[gPID]['purpose']}", level='error', caller_id=caller_id)
            return False, []
        
        # now check that the filepaths make sense
        filepath_purpose = exp.plate_location_sample[gPID].get('filepath_purpose', None)
        if not isinstance(filepath_purpose, dict) or len(filepath_purpose) == 0:
            exp.plate_location_sample[gPID]['filepath_purpose'] = {fp:'primer_volume'}
        else:
            if fp in filepath_purpose:
                if filepath_purpose[fp] == 'primer_volume':
                    if not overwrite_plates:
                        m(f'Primer volume plate already found in experiment, overwriting not allowed', 
                                level='failure', caller_id=caller_id)
                        return False, []
                    m(f'replacing existing primer volume for {gPID}, wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_volume'}}
                      
                elif filepath_purpose[fp] not in ['primer_layout', 'primer_volume']:
                    m(f'existing file of path {fp} has new purpose "primer_volume" instead of {filepath_purpose[fp]}, '+ 
                            f'wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_volume'}}          
    else:
        exp.plate_location_sample[gPID] = {'purpose': 'primer', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'primer_volume'}}
                      
    for well in well_records[gPID]:
        if well not in exp.plate_location_sample[gPID]:
            exp.plate_location_sample[gPID]['wells'].add(well)
            exp.plate_location_sample[gPID][well] = {}
        exp.plate_location_sample[gPID][well]['volume'] = well_records[gPID][well]['volume']
                    
    m(f"added primer volumes from {fp}", level='end')
    return True, [gPID]

 
def parse_index_layout(exp, fp, user_gpid=None, caller_id=None, overwrite_plates=True):
    """ 
    Add index plates with well and index columns 
    Inputs: experiment, fp - filepath usually containing the PID, user_gpid is an override for the filepath
    Non-sample plates can be loaded straight into experiment.plate_location_sample
    Returns True on success
    caller_id (str) is the name of the associated display unit for user messages
    """
    if user_gpid:
        PID = util.unguard_pbc(user_gpid, silent=True)
        gPID = util.guard_pbc(user_gpid, silent=True)
    else:
        PID = Path(fp).name.split('_')[0]  # assumes first field is PID
        gPID = util.guard_pbc(PID, silent=True)   
    
    # parse the file first to make sure it works
    well_records = {gPID:{}}
    with open(fp, 'rt', errors='ignore') as data:
        for i, row in enumerate(csv.reader(data, delimiter=',', quoting=csv.QUOTE_MINIMAL)):
            if i == 0 or row == '':
                continue  # header or blank
            try:
                well = util.unpadwell(row[0])
            except UnboundLocalError as exc:
                m(f'Did you upload a volume file by mistake? {exc}', level='error', caller_id=caller_id)
                return False, []
            if well in well_records[gPID]:
                m(f'skipping duplicate well entry {well} in {fp}', level='warning', caller_id=caller_id)
                continue
            well_records[gPID][well] = {}
            idt_name = row[1]
            index = row[2]
            bc_name = row[3]
            oligo = row[4]
            well_records[gPID][well]['idt_name'] = idt_name
            well_records[gPID][well]['index'] = index
            well_records[gPID][well]['bc_name'] = bc_name
            well_records[gPID][well]['oligo'] = oligo

    # check if the plate already exists and that its purpose is related but not radically different
    if gPID in exp.plate_location_sample:
        plate_purpose = exp.plate_location_sample[gPID].get('purpose', None) 
        if plate_purpose is None:
            # this should not happen, every plate has a purpose. Wipe the existing plate
            exp.delete_plate(gPID)
            exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                    'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_layout'}}
            m(f"Creating new index plate record for {PID}", level='info', caller_id=caller_id)
        elif plate_purpose != 'index':
            m(f"index plate PID: {PID} matches existing plate entry of different purpose "+\
                    f"{exp.plate_location_sample[gPID]['purpose']}", level='error', caller_id=caller_id)
            return False, []
        
        # now check that the filepaths make sense
        filepath_purpose = exp.plate_location_sample[gPID].get('filepath_purpose', None)
        if not isinstance(filepath_purpose, dict) or len(filepath_purpose) == 0:
            exp.plate_location_sample[gPID]['filepath_purpose'] = {fp:'index_layout'}
        else:
            if fp in filepath_purpose:
                if filepath_purpose[fp] == 'index_layout':
                    if not overwrite_plates:
                        m(f'Index plate already found in experiment, overwriting not allowed', 
                                level='failure', caller_id=caller_id)
                        return False, []
                    m(f'replacing existing index layout for {gPID}, wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_layout'}}
                      
                elif filepath_purpose[fp] not in ['index_layout', 'index_volume']:
                    m(f'existing file of path {fp} has new purpose "index_layout" instead of {filepath_purpose[fp]}, '+ 
                            f'wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_layout'}}
    else:
        exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_layout'}}
                    
    for well in well_records[gPID]:
        if well not in exp.plate_location_sample[gPID]:
            exp.plate_location_sample[gPID]['wells'].add(well)
            exp.plate_location_sample[gPID][well] = {}
        exp.plate_location_sample[gPID][well]['idt_name'] = well_records[gPID][well]['idt_name']
        exp.plate_location_sample[gPID][well]['index'] = well_records[gPID][well]['index']
        exp.plate_location_sample[gPID][well]['bc_name'] = well_records[gPID][well]['bc_name']
        exp.plate_location_sample[gPID][well]['oligo'] = well_records[gPID][well]['oligo']

    m(f"added index plate layouts from {fp}", level='end')        
    return True, [gPID]
    

def parse_index_volume(exp, fp, user_gpid=None, caller_id=None, overwrite_plates=True):
    """ add volume information to a single barcode plate, with a matched, possibly-existing plate name.
    To make life fun there are two possible formats:
    Format 1: plate layout (generated through Echo test software)
    Format 2: column layout (generated through Echo main interface)
        
    Inputs: experiment, fp - filepath usually containing the PID, user_gpid is an override for the filepath
    Non-sample plates can be loaded straight into experiment.plate_location_sample
    Returns True on success
    caller_id (str) is the name of the associated display unit for user messages
    """
    if user_gpid:
        PID = util.unguard_pbc(user_gpid, silent=True)
        gPID = user_gpid
    else:
        PID = Path(fp).name.split('_')[0]  # assumes first field is PID
        gPID = util.guard_pbc(PID, silent=True)   
    
    # parse the file first to make sure it works
    well_records = {gPID:{}}

    plate_format = False
    hdr_seen = False
    with open(fp, 'rt', errors='ignore') as data:
        for i, row in enumerate(csv.reader(data, delimiter=',', quoting=csv.QUOTE_MINIMAL)):
            if i == 0:
                if row[0].lower().startswith('date'):
                    plate_format = True
                continue
            if plate_format:
                # format 1 - volume in r,c plate matrix
                # matrix format - 2 initial lines, a blank line then a matrix
                #next(src), next(src) # skip two more lines
                if not row or row[0].strip() == '':
                    continue
                elif row[0].strip() == 'A':
                    hdr_seen = True
                elif not hdr_seen:
                    continue
                    
                plate_row = row[0].strip()
                for col, v in enumerate(row):
                    if col > 0 and v.strip() != '':
                        well = plate_row+str(col)
                        if well in well_records[gPID]:
                            m(f'duplicate well entry {well} seen in plate {PID}, overwriting', level='warning', caller_id=caller_id)
                        well_records[gPID][well] = {'volume': float(v)*1000}  
            else:                
                #format 2 - one row per well - well ID in r[3], volume in r[5] or r[6]
                #if i == 0 and line.startswith('Run ID'):
                if row.startswith('[DETAILS]'):
                    continue
                if ''.join(row).strip() == '':
                    continue
                if not hdr_seen:
                    hdr_seen = True
                    continue
                
                # finally, the data    
                well = util.unpadwell(row[3])
                if well in well_records[gPID]:
                    m(f'duplicate well entry {well} seen in plate {PID}, overwriting', level='warning', caller_id=caller_id)
                well_records[gPID][well] = {'volume': int(float(row[5]))*1000}

    # check if the plate already exists and that its purpose is related but not radically different
    if gPID in exp.plate_location_sample:
        plate_purpose = exp.plate_location_sample[gPID].get('purpose', None) 
        if plate_purpose is None:
            # this should not happen, every plate has a purpose. Wipe the existing plate
            exp.delete_plate(gPID)
            exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                    'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_volume'}}
            m(f"Creating new index plate record for {PID}", level='info', caller_id=caller_id)
        elif plate_purpose != 'index':
            m(f"index plate PID: {PID} matches existing plate entry of different purpose "+\
                    f"{exp.plate_location_sample[gPID]['purpose']}", level='error', caller_id=caller_id)
            return False, []

        # now check that the filepaths make sense
        filepath_purpose = exp.plate_location_sample[gPID].get('filepath_purpose', None)
        if not isinstance(filepath_purpose, dict) or len(filepath_purpose) == 0:
            exp.plate_location_sample[gPID]['filepath_purpose'] = {fp:'index_volume'}
        else:
            if fp in filepath_purpose:
                if filepath_purpose[fp] == 'index_volume':
                    if not overwrite_plates:
                        m(f'index plate already found in experiment, overwriting not allowed', 
                                level='failure', caller_id=caller_id)
                        return False, []
                    m(f'replacing existing index volumes for {gPID}, wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_volume'}}
                      
                elif filepath_purpose[fp] not in ['index_layout', 'index_volume']:
                    m(f'existing file of path {fp} has new purpose "index_volume" instead of {filepath_purpose[fp]}, '+ 
                            f'wiping plate information', level='warning', caller_id=caller_id)
                    exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                            'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_volume'}}
    else:
        exp.plate_location_sample[gPID] = {'purpose': 'index', 'source':'user', 'wells':set(), 
                'plate_type':util.PLATE_TYPES['Echo384'],'filepath_purpose':{fp:'index_volume'}}
                    
    for well in well_records[gPID]:
        if well not in exp.plate_location_sample[gPID]:
            exp.plate_location_sample[gPID]['wells'].add(well)
            exp.plate_location_sample[gPID][well] = {}
        exp.plate_location_sample[gPID][well]['volume'] = well_records[gPID][well]['volume']
        
    m(f'completed parsing index volumes {fp}', level='end')
    
    return True, [gPID]


def parse_assay_primer_map(exp, fp, caller_id=None, overwrite_plates=True):
    """ mapping of assay name, assay family name, and primers
    The assaylist file has two fixed columns: primer and assay
        - every subsequent column is for an extra primer in that assay family
    caller_id (str) is the name of the associated display unit for user messages
    In Experiment we maintain:
        self.assayfam_primer = {}  # mapping of assay families to list of primers
        self.assay_assayfam = {}  # mapping of assay to assay family
        self.assayfam_assay = {}  # mapping of assay family to list of assays, possibly redundant
        self.primer_assayfam = {}  # reverse mapping
        
    Case is strictly enforced - we don't bother even checking case mistake, just reporting them
    """
    m(f'parsing assay primer map {fp}', level='begin')
    # read the whole file before adding entries
    local_assayfam_primers = {}
    local_assay_assayfam = {}
    local_assayfam_assays = {}
    local_primer_assayfam = {} 
    with open(fp, 'rt', errors='ignore') as data:
        for i, row in enumerate(csv.reader(data, delimiter=',', quoting=csv.QUOTE_ALL)):
            if len(row) < 2:
                continue  # empty row
            if i == 0:
                continue  # header
            if row == '' or row[0] == '' or row[1] == '':
                continue  # empty row

            assay = row[0].strip()
            if not assay:
                continue  # just skip it
            assay_fam = row[1].strip()
            if not assay_fam:
                m(f'no assay family associated with {assay} in row {i+1}', level='warning', caller_id=caller_id)
                continue
            primers = [p.strip() for p in row[2:] if p.strip() != '']
            if not primers:
                m(f'no primers associated with assay {assay} in row {i+1}', level='warning', caller_id=caller_id)
                continue

            # assay to assay family
            if assay in local_assay_assayfam:
                m(f'{assay} has duplicate entry {local_assay_assayfam[assay]} seen in row {i+1}', 
                        level='warning', caller_id=caller_id)
            local_assay_assayfam[assay] = assay_fam

            # assay family to assay
            if assay_fam not in local_assayfam_assays:
                local_assayfam_assays[assay_fam] = []
            local_assayfam_assays[assay_fam].append(assay)

            # assay family to primers
            if assay_fam not in local_assayfam_primers:
                local_assayfam_primers[assay_fam] = primers
            else:
                for primer in primers:
                    if primer not in local_assayfam_primers[assay_fam]:
                        local_assayfam_primers[assay_fam].append(primer)

            # reverse mapping, primers to assayfams
            for primer in primers:
                if primer in local_primer_assayfam:
                    if local_primer_assayfam[primer] != assay_fam:
                        msg = f'primer {primer} has multiple assayfams: '+\
                                f'{local_primer_assayfam[primer]} and {assay_fam}'
                        m(msg, level='warning', caller_id=caller_id)
                        continue
                local_primer_assayfam[primer] = assay_fam

    exp.assay_assayfam = local_assay_assayfam.copy()
    exp.assayfam_assays = local_assayfam_assays.copy()
    exp.assayfam_primers = local_assayfam_primers.copy()
    exp.primer_assayfam = local_primer_assayfam.copy()
    
    m(f'added assay/assayfam/primer mapping from {fp}', level='end')
    return True


def parse_reference_sequences(exp, fp, purpose, caller_id=None, overwrite_plates=True):
    """
    read in reference (target) IDs and sequences.
    May be added when pipeline is locked.
    Any number of reference files are supported
    purpose may be 'rodentity_reference', 'custom_reference' or 'amplicon_reference'
    Saved as exp.reference_sequences[(fp, purpose)] = [(ref_id,sequence)]
    caller_id (str) is the name of the associated display unit for user messages
    Note: In version 1.02.002 and earlier only one reference file was allowed
        and it was saved as exp.reference_sequences[fp] = {ref_id:sequence}
    """
    m(f'parsing reference sequences {fp}', level='begin')
    sequences = parse_fasta(open(fp, 'rb'), caller_id=caller_id)
    if not sequences:
        m(f'could not read reference seq file {fp}', level='error', caller_id=caller_id)
        return False
    exp.reference_sequences[(fp,purpose)] = [(ref,sequences[ref]) for ref in sequences]
    m(f'{len(exp.reference_sequences[(fp,purpose)])} reference sequences imported from {fp}', level='info', caller_id=caller_id)
    return True
  

def parse_dna_plate(exp, filepath, caller_id=None, overwrite_plates=True):
    """
    Filepath should point to an Echo_384_COC file/Nimbus output/Echo input
    caller_id (str) is the name of the associated display unit for user messages
    """
    m(f'loading DNA plates {filepath}', level='begin')
    # "Echo_384_COC_0001_" + ug_dnaBC + "_0.csv"
    with open(filepath, 'rt') as f:
        #RecordId	TRackBC	TLabwareId	TPositionId	SRackBC	SLabwareId	SPositionId
        #1	p2021120604p	Echo_384_COC_0001	A1	p2111267p	ABg_96_PCR_NoSkirt_0001	A1
        gPID = util.guard_pbc(filepath.split('_')[-2], silent=True)
        exp.plate_location_sample[gPID] = {'purpose':'dna','wells':set(),'source':'',
                'plate_type':util.PLATE_TYPES['Echo384']}
        source_plate_set = set()
        for i, line in enumerate(f):
            if i == 0:  # header
                # get named columns
                cols = [c.strip() for c in line.split(',')]
                source_plate_bc_col = [j for j,c in enumerate(cols) if c=='"SRackBC"'][0]
                source_well_col = [j for j,c in enumerate(cols) if c=='"SPositionId"'][0]
                source_bc_col = [j for j,c in enumerate(cols) if c=='"SPositionBC"'][0]
                dest_plate_bc_col = [j for j,c in enumerate(cols) if c=='"TRackBC"'][0]
                dest_well_col = [j for j,c in enumerate(cols) if c=='"TPositionId"'][0]
                continue
            if line.strip() == '':
                continue  # skip empty lines
            cols = [c.strip().strip('\"') for c in line.split(',')]
            source_bc = cols[source_bc_col]
            if source_bc == '0M':
                continue  # empty well
            source_pos = util.unpadwell(cols[source_well_col])
            source_plate = util.guard_pbc(cols[source_plate_bc_col], silent=True)
            source_plate_set.add(source_plate)
            dest_plate = util.guard_pbc(cols[dest_plate_bc_col], silent=True)
            dest_pos = util.unpadwell(cols[dest_well_col])
            if dest_plate != gPID:
                m(f"{gPID} doesn't match {dest_plate} as declared in Echo_384_COC file: {filepath}",
                        level='error', caller_id=caller_id)
                return False, [gPID]         
            try:
                exp.plate_location_sample[gPID][dest_pos] = exp.plate_location_sample[source_plate][source_pos]
            except:
                m(f"cannot locate {gPID=} {dest_pos=} {source_plate=} {source_pos=}", 
                        level='critical', caller_id=caller_id)
                return False, [gPID]
            exp.plate_location_sample[gPID]['wells'].add(dest_pos)
            exp.plate_location_sample[gPID]['samplePlate'] = source_plate
            exp.plate_location_sample[gPID]['sampleWell'] = source_pos
        exp.plate_location_sample[gPID]['source'] = ','.join(source_plate_set)
    return True, [gPID]


def myopen(fn):
    """ Bob's function to handle gzip transparently """
    if fn.endswith('.gz') :
        return gzip.open(fn, "rt", errors='ignore')
    return open(fn, errors="ignore")


def read_html_template(html_fn):
    """ 
    Read an html file from the library folder, ready for additional fields 
    Used by cgi-dashboard.py and makehtml.py
    template uses {!field!} instead of python format fields - easier to edit.
    Files should use the .tpl (template) extension as they are not strictly html
    """
    tfn = os.path.join('..', 'library', html_fn)
    with open(tfn, errors='ignore') as src:
        htmlcode = src.read()
    # template uses {!field!} instead of python format fields - easier to edit.
    htmlfmt = htmlcode.replace('{', '{{').replace('}', '}}').replace('{{!', '{').replace('!}}', '}')
    return htmlfmt


def read_csv_or_excel_from_stream(multi_stream):
    """
    Take a stream of one or more table files from CSV or Excel and convert to comma delimited rows
    return a list of names and a list of tables

    DEPRECATED
    """
    manifest_names = []
    manifest_tables = []
    for file_stream in multi_stream:
        manifest_name = file_stream.name
        #print(f'{manifest_name=}', file=sys.stderr)
        if manifest_name.lower().endswith('xlsx'):
            workbook = openpyxl.load_workbook(BytesIO(file_stream.getvalue()))
            sheet = workbook.active
            rows = [','.join(map(str,cells)) for cells in sheet.iter_rows(values_only=True)]
            #print(rows, file=sys.stderr)
        else:
            rows = StringIO(file_stream.getvalue().decode("utf-8"))
        manifest_names.append(manifest_name)
        manifest_tables.append(rows)
    return manifest_names, manifest_tables


def check_non_ascii(file_content, issue_num):
    """
    A function to check for non-ASCII characters.
    Inputs: character stream
    Outputs: list of dictionaries
    """
    issues = []
    lines = file_content.splitlines()
    line_num = 1  # Start with the first line

    for line_num, line in enumerate(lines):
        for idx, char in enumerate(line):
            if ord(char) > 127:  # ASCII characters have values from 0 to 127
                issue_num += 1
                issues.append({
                    'Issue Number': issue_num,
                    'Line Number': line_num,
                    'Issue': f"Invalid character in sequence header: '{char}' at position {idx + 1}"
                })
        line_num += 1  # Move to the next line number
    return issues  


def check_valid_sequence(file_content, issue_num):
    """
    Function to check that sequences contain only A, T, C, G, N, (, ), [, ]
        brackets and parentheses are allowed for reference sequences to denote variable regions
    Spaces and tabs will not be reported here, only checked in check_gaps
    Inputs: character stream
    Outputs: list of dictionaries
    """
    issues = []
    lines = file_content.splitlines()
    line_num = 1  # Start with the first line

    for line in lines:
        if line.startswith(">") or not line.strip():  # Ignore headers and blank lines
            line_num += 1
            continue
        
        sequence = line.strip().replace(" ", "").replace("\t", "")  # Clean the sequence

        for idx, char in enumerate(sequence):
            if char not in "AaTtCcGgNn()[]":  # Check for valid characters
                issue_num += 1
                issues.append({
                    'Issue Number': issue_num,
                    'Line Number': line_num,
                    'Issue': f"Invalid character in sequence: '{char}' at position {idx + 1}"
                })
        
        line_num += 1  # Move to the next line number

    return issues


def check_gaps(file_content, issue_num):
    """
    Function to check for gaps (spaces or tabs) within sequences.
    Inputs: character stream
    Outputs: list of dictionaries
    """
    issues = []
    lines = file_content.splitlines()
    line_num = 1  # Start with the first line

    for line in lines:
        if line.startswith(">") or not line.strip():  # Ignore headers and blank lines
            line_num += 1
            continue

        if " " in line or "\t" in line:
            issue_num += 1
            issues.append({
                'Issue Number': issue_num,
                'Line Number': line_num,
                'Issue': "Gap (space or tab) in sequence"
            })
        
        line_num += 1  # Move to the next line number

    return issues

### FASTA parsing/checking funtions below ###

def check_blank_lines(file_content, issue_num):
    """
    FASTA files should not contain blank lines
    Inputs: character stream
    Outputs: issues as a list of dictionaries
    """
    issues = []
    lines = file_content.splitlines()
    line_num = 1  # Start with the first line

    for line in lines:
        if not line.strip():  # Blank line
            issue_num += 1
            issues.append({
                'Issue Number': issue_num,
                'Line Number': line_num,
                'Issue': "Blank line found"
            })
        line_num += 1  # Move to the next line number
    return issues


def check_fasta_file(file_content):
    """
    Issue feature checks for FASTA specific file features
    Inputs: character stream
    Outputs: issues as a list of dictionaries
    """
    issues = []
    issues.extend(check_non_ascii(file_content, len(issues)))
    issues.extend(check_valid_sequence(file_content, len(issues)))
    issues.extend(check_gaps(file_content, len(issues)))
    issues.extend(check_blank_lines(file_content, len(issues)))
    return issues   


def _report_issues(issues, caller_id=None):
    """
    Helper: report issues in a list of dictionaries
    """
    if issues:
        for issue in issues:
            m(f"Issue {issue['Issue Number']}: {issue['Issue']} at line {issue['Line Number']}", level='error', caller_id=caller_id)


def read_text_file(file_stream, caller_id=None):
    """
    Common text file handling to search for character issues before parsing
    """
    # Read the raw file
    rawfile = file_stream.read()
    
    # Detect encoding and Decode the file content
    result = chardet.detect(rawfile)
    if not result:
        m('Could not interpret file format', level='error',caller_id=caller_id)
        return None
    charenc = result['encoding']
    if not charenc:
        m('Could not interpret file format', level='error', caller_id=caller_id)
        return None

    file_content = rawfile.decode(charenc)
    
    # Check for issues in the file content
    issues = check_non_ascii(file_content,0)
    if issues:
        _report_issues(issues)
        return None
    else:
        return file_content
    

def parse_fasta(file_stream, caller_id=None):
    """
    Parse a FASTA file and return a dictionary of sequences
    """
    file_content = read_text_file(file_stream, caller_id=caller_id)
    if not file_content:
        return None
    # Check for issues in the file content
    issues = check_fasta_file(file_content)
    if issues:
        _report_issues(issues)
        return None
    else:
        # Parse the FASTA file
        sequences = {}
        lines = file_content.splitlines()
        seq_id = None
        seq = []
        for line in lines:
            if line.startswith(">"):
                if seq_id:
                    sequences[seq_id] = "".join(seq)
                seq_id = line[1:].strip()
                seq = []
            else:
                seq.append(line.strip())
        if seq_id:
            sequences[seq_id] = "".join(seq)
        return sequences


if __name__ == '__main__':
    """ library only """
    pass